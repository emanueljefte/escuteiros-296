from openpyxl import load_workbook
from supabase import create_client
import os

def handler(request):
    supabase = create_client(os.environ["SUPABASE_URL"], os.environ["SUPABASE_SERVICE_KEY"])
    registos = supabase.table("escuteiros").select("*").execute().data

    wb = load_workbook("template/dashboard_template.xlsx")
    ws = wb["Dados"]

    for i, r in enumerate(registos, start=2):
        ws.cell(row=i, column=1, value=r.get("numero_inscricao"))
        ws.cell(row=i, column=2, value=r.get("nome_completo"))
        # ... resto das colunas, mesma ordem dos headers
        ws.cell(row=i, column=27, value="Sim" if r.get("doenca") else "Não")
        ws.cell(row=i, column=29, value="Sim" if r.get("alergia") else "Não")
        # etc.

    caminho = "/tmp/Cadastramento_Agrupamento_296.xlsx"
    wb.save(caminho)
    return caminho  # devolvido como download